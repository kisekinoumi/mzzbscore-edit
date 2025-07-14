"""
Excel处理服务

提供Excel文件读取、写入和样式处理的服务。
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Border, Side
import shutil
import os
import tempfile
from typing import Dict, Any, Optional, Tuple

from app.core.base import BaseService, IFileHandler
from app.models.data_models import RankingResult, ProcessingConfig
from app.config.constants import ExcelColumns, EXCEL_STYLE_CONFIG
from app.utils.exceptions import ExcelProcessingError, FileOperationError, DataFormatError
from app.utils.validators import FileValidator, DataValidator
from app.utils.logger import Logger


class ExcelService(BaseService, IFileHandler):
    """Excel文件处理服务"""
    
    def __init__(self, logger: Optional[Logger] = None):
        """
        初始化Excel服务
        
        Args:
            logger: 日志记录器
        """
        super().__init__(logger.get_logger("ExcelService") if logger else None)
        self._style_config = EXCEL_STYLE_CONFIG
        self._alignment = Alignment(horizontal='left', vertical='center')
        self._border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def initialize(self) -> bool:
        """
        初始化服务
        
        Returns:
            bool: 初始化是否成功
        """
        try:
            self.logger.info("ExcelService初始化开始")
            
            # 验证依赖项
            try:
                import pandas as pd
                import openpyxl
                self.logger.debug("依赖项验证通过")
            except ImportError as e:
                raise ExcelProcessingError(f"缺少必要的依赖项: {e}")
            
            self._set_initialized(True)
            self.logger.info("ExcelService初始化成功")
            return True
            
        except Exception as e:
            self.logger.error(f"ExcelService初始化失败: {e}")
            return False
    
    def validate_file(self, file_path: str) -> bool:
        """
        验证文件
        
        Args:
            file_path: 文件路径
            
        Returns:
            bool: 文件是否有效
        """
        try:
            FileValidator.validate_file_exists(file_path)
            FileValidator.validate_file_readable(file_path)
            FileValidator.validate_excel_file(file_path)
            return True
        except Exception as e:
            self.logger.error(f"文件验证失败: {e}")
            raise
    
    def read_file(self, file_path: str) -> pd.DataFrame:
        """
        读取Excel文件
        
        Args:
            file_path: 文件路径
            
        Returns:
            pd.DataFrame: 读取的数据
        """
        wb = None
        try:
            self.logger.info(f"开始读取Excel文件: {file_path}")
            
            # 验证文件
            self.validate_file(file_path)
            
            # 使用openpyxl读取数据
            try:
                wb = openpyxl.load_workbook(file_path)
            except openpyxl.utils.exceptions.InvalidFileException as e:
                raise DataFormatError(f"无效的Excel文件: {e}")
            
            if not wb.worksheets:
                raise DataFormatError("Excel文件中没有工作表")
                
            ws = wb.active
            if ws is None:
                raise DataFormatError("无法获取活动工作表")
            
            # 检查工作表结构
            if ws.max_row < 3:
                raise DataFormatError("Excel文件至少需要3行数据（标题行、表头行、数据行）")
            
            # 读取表头（第2行）
            headers = []
            valid_cols = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=2, column=col).value
                if cell_value is not None and str(cell_value).strip():
                    headers.append(str(cell_value).strip())
                    valid_cols.append(col)
            
            if not headers:
                raise DataFormatError("没有找到有效的列标题")
            
            # 读取数据行（从第3行开始）
            data_rows = []
            for row in range(3, ws.max_row + 1):
                row_data = []
                for col in valid_cols:
                    try:
                        cell_value = ws.cell(row=row, column=col).value
                        row_data.append(cell_value)
                    except Exception as e:
                        self.logger.warning(f"读取单元格 ({row}, {col}) 时发生错误: {e}")
                        row_data.append(None)
                data_rows.append(row_data)
            
            # 创建DataFrame
            df = pd.DataFrame(data_rows, columns=headers)
            
            # 数据清洗
            original_rows = len(df)
            key_col = ExcelColumns.ORIGINAL_NAME
            
            if key_col not in df.columns:
                raise DataFormatError(f"Excel文件中缺少必需的列: '{key_col}'")
            
            # 删除空值行
            df.dropna(subset=[key_col], inplace=True)
            df = df[df[key_col].astype(str).str.strip() != '']
            
            cleaned_rows = len(df)
            
            if cleaned_rows == 0:
                raise DataFormatError("清洗后没有有效数据")
            
            if original_rows > cleaned_rows:
                self.logger.info(f"数据清洗: 已忽略 {original_rows - cleaned_rows} 条无效行")
            
            self.logger.info(f"Excel文件读取成功，共 {cleaned_rows} 条有效记录")
            return df
            
        except (FileOperationError, DataFormatError):
            raise
        except Exception as e:
            raise ExcelProcessingError(f"读取Excel文件时发生错误: {e}")
        finally:
            if wb:
                try:
                    wb.close()
                except:
                    pass
    
    def write_file(self, file_path: str, data: Any) -> bool:
        """
        写入Excel文件
        
        Args:
            file_path: 文件路径
            data: 要写入的数据
            
        Returns:
            bool: 写入是否成功
        """
        if isinstance(data, RankingResult):
            return self._write_ranking_result(file_path, data)
        else:
            raise ExcelProcessingError("不支持的数据类型")
    
    def _write_ranking_result(self, file_path: str, result: RankingResult) -> bool:
        """
        写入排名结果到Excel文件
        
        Args:
            file_path: 输出文件路径
            result: 排名结果
            
        Returns:
            bool: 写入是否成功
        """
        temp_file = None
        wb = None
        
        try:
            self.logger.info(f"开始写入排名结果到: {file_path}")
            
            # 验证输出目录
            FileValidator.validate_output_directory(file_path)
            
            # 获取输入文件路径（从result中获取或使用默认值）
            input_file = getattr(result, 'input_file', 'mzzb.xlsx')
            
            # 创建临时文件
            output_dir = os.path.dirname(file_path) or "."
            temp_file = tempfile.NamedTemporaryFile(
                delete=False, 
                suffix='.xlsx', 
                dir=output_dir
            )
            temp_file.close()
            temp_path = temp_file.name
            
            # 复制输入文件到临时文件
            shutil.copy2(input_file, temp_path)
            
            # 打开工作簿
            wb = openpyxl.load_workbook(temp_path)
            ws = wb.active
            header_row = 2
            
            # 收集原始超链接
            original_hyperlinks = self._collect_hyperlinks(ws)
            
            # 插入排名列
            column_offset = self._insert_ranking_columns(ws, result, header_row)
            
            # 重新应用超链接
            self._reapply_hyperlinks(ws, original_hyperlinks, column_offset)
            
            # 写入数据
            final_col_map = {cell.value: cell.column for cell in ws[header_row] if cell.value}
            self._write_data_to_worksheet(ws, result, final_col_map, header_row)
            
            # 应用样式
            if hasattr(result, 'apply_styles') and result.apply_styles:
                self._apply_column_styles(ws, final_col_map, header_row)
            
            # 保存文件
            wb.save(temp_path)
            wb.close()
            wb = None
            
            # 原子性移动文件
            self._atomic_move_file(temp_path, file_path)
            temp_file = None
            
            self.logger.info(f"排名结果写入成功: {file_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"写入排名结果失败: {e}")
            raise ExcelProcessingError(f"写入文件失败: {e}")
        finally:
            # 清理资源
            if wb:
                try:
                    wb.close()
                except:
                    pass
            
            if temp_file and os.path.exists(temp_file.name):
                try:
                    os.unlink(temp_file.name)
                except:
                    pass
    
    def _collect_hyperlinks(self, ws) -> Dict[Tuple[int, int], Dict[str, Any]]:
        """收集工作表中的超链接"""
        hyperlinks = {}
        try:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.hyperlink:
                        hyperlinks[(cell.row, cell.column)] = {
                            'hyperlink': cell.hyperlink,
                            'value': cell.value
                        }
        except Exception as e:
            self.logger.warning(f"收集超链接时发生错误: {e}")
        return hyperlinks
    
    def _insert_ranking_columns(self, ws, result: RankingResult, header_row: int) -> Dict[int, int]:
        """插入排名列并返回列偏移映射"""
        try:
            self.logger.info("开始检查和插入排名列")
            column_offset = {}
            
            # 获取当前所有列的映射 {列名: 列索引}
            current_columns = {}
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=header_row, column=col_idx).value
                if cell_value:
                    current_columns[str(cell_value).strip()] = col_idx
            
            self.logger.debug(f"当前列映射: {current_columns}")
            
            # 需要处理的排名列配置: (评分列名, total列名, 排名列名)
            ranking_configs = [
                ("Bangumi", "Bangumi_total", "Bangumi_Rank"),
                ("Anilist", "Anilist_total", "Anilist_Rank"), 
                ("MyAnimelist", "MyAnimelist_total", "Myanimelist_Rank"),
                ("Filmarks", "Filmarks_total", "Filmarks_Rank")
            ]
            
            total_inserted = 0
            
            for score_col, total_col, rank_col in ranking_configs:
                try:
                    # 检查排名列是否已存在
                    if rank_col in current_columns:
                        self.logger.debug(f"排名列 {rank_col} 已存在于第 {current_columns[rank_col]} 列")
                        continue
                    
                    # 查找 total 列的位置
                    if total_col not in current_columns:
                        self.logger.warning(f"找不到 {total_col} 列，跳过 {rank_col} 的插入")
                        continue
                    
                    total_col_idx = current_columns[total_col]
                    insert_position = total_col_idx + 1
                    
                    self.logger.debug(f"将在第 {insert_position} 列插入 {rank_col}")
                    
                    # 在 total 列右侧插入新列
                    ws.insert_cols(insert_position)
                    
                    # 设置列标题
                    ws.cell(row=header_row, column=insert_position).value = rank_col
                    
                    # 更新列映射（所有在插入位置右侧的列都要向右偏移1）
                    new_current_columns = {}
                    for col_name, col_idx in current_columns.items():
                        if col_idx >= insert_position:
                            new_current_columns[col_name] = col_idx + 1
                        else:
                            new_current_columns[col_name] = col_idx
                    
                    # 添加新插入的列
                    new_current_columns[rank_col] = insert_position
                    current_columns = new_current_columns
                    
                    total_inserted += 1
                    self.logger.debug(f"成功插入排名列: {rank_col} 在第 {insert_position} 列")
                    
                except Exception as e:
                    self.logger.error(f"插入排名列 {rank_col} 失败: {e}")
            
            # 现在填入排名数据
            # 不再需要在这里处理排名数据，因为排名服务已经正确设置了"NaN"文本
            # if hasattr(result, 'valid_data') and result.valid_data is not None:
            #     self._fill_ranking_data(ws, result, current_columns, header_row)
            
            self.logger.info(f"排名列处理完成，插入了 {total_inserted} 列")
            return column_offset
            
        except Exception as e:
            self.logger.error(f"处理排名列时发生错误: {e}")
            return {}
    
    def _fill_ranking_data(self, ws, result: RankingResult, column_mapping: Dict[str, int], header_row: int):
        """填入排名数据到对应的列"""
        try:
            self.logger.debug("开始填入排名数据")
            
            # 创建有效数据字典便于查找
            valid_data_dict = result.valid_data.set_index("原名").to_dict('index')
            
            # 排名列列表
            rank_columns = ["Bangumi_Rank", "Anilist_Rank", "Myanimelist_Rank", "Filmarks_Rank"]
            
            # 统计写入情况
            write_stats = {col: {"numeric": 0, "nan_text": 0, "missing": 0} for col in rank_columns}
            
            # 写入排名数据到对应行
            rows_written = 0
            for row_idx in range(header_row + 1, ws.max_row + 1):
                # 获取动漫名称（第1列）
                anime_name_cell = ws.cell(row=row_idx, column=1)
                anime_name = anime_name_cell.value
                
                if anime_name and anime_name in valid_data_dict:
                    # 这是有效条目（计算综合评分的条目）
                    source_data = valid_data_dict[anime_name]
                    
                    # 为每个排名列写入数据
                    for rank_col in rank_columns:
                        if rank_col in column_mapping:
                            col_idx = column_mapping[rank_col]
                            
                            if rank_col in source_data:
                                rank_value = source_data[rank_col]
                                # 使用更严格的检查：既要检查pd.notna又要检查不是pd.NA
                                if pd.notna(rank_value) and rank_value is not pd.NA:
                                    # 有排名数据，写入数值
                                    ws.cell(row=row_idx, column=col_idx).value = rank_value
                                    write_stats[rank_col]["numeric"] += 1
                                else:
                                    # 有效条目但该站点没有排名数据（rank_value是pd.NA或NaN），写入"NaN"文本
                                    ws.cell(row=row_idx, column=col_idx).value = "NaN"
                                    write_stats[rank_col]["nan_text"] += 1
                                    self.logger.debug(f"为有效条目 '{anime_name}' 的 {rank_col} 写入'NaN'文本（原值: {rank_value}）")
                            else:
                                # 有效条目但没有这个排名数据列，写入"NaN"文本
                                ws.cell(row=row_idx, column=col_idx).value = "NaN"
                                write_stats[rank_col]["missing"] += 1
                    
                    rows_written += 1
                else:
                    # 非有效条目（不计算综合评分的条目），不处理排名数据
                    pass
            
            # 输出统计信息
            self.logger.debug(f"排名数据填入完成，处理了 {rows_written} 行有效条目")
            for col, stats in write_stats.items():
                if stats["numeric"] + stats["nan_text"] + stats["missing"] > 0:
                    self.logger.info(f"{col}: {stats['numeric']}个数字排名, {stats['nan_text']}个'NaN'文本, {stats['missing']}个缺失列")
            
        except Exception as e:
            self.logger.error(f"填入排名数据时发生错误: {e}")
            import traceback
            self.logger.error(traceback.format_exc())
    
    def _reapply_hyperlinks(self, ws, hyperlinks: Dict, column_offset: Dict[int, int]):
        """重新应用超链接"""
        try:
            self.logger.debug("开始重新应用超链接")
            
            # 首先清除所有超链接
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.hyperlink:
                        cell.hyperlink = None
            
            # 获取原始文件的列映射（从超链接收集时的信息）
            original_columns = {}
            # 重新读取原始文件获取列映射
            try:
                temp_original = openpyxl.load_workbook("mzzb.xlsx")
                temp_ws = temp_original.active
                for col_idx in range(1, temp_ws.max_column + 1):
                    cell_value = temp_ws.cell(row=2, column=col_idx).value
                    if cell_value:
                        original_columns[col_idx] = str(cell_value).strip()
                temp_original.close()
                self.logger.debug(f"原始列映射: {original_columns}")
            except Exception as e:
                self.logger.error(f"无法读取原始文件列映射: {e}")
                return
            
            # 获取当前文件的列映射
            current_columns = {}
            column_name_to_index = {}
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=2, column=col_idx).value
                if cell_value:
                    col_name = str(cell_value).strip()
                    current_columns[col_idx] = col_name
                    column_name_to_index[col_name] = col_idx
            
            self.logger.debug(f"当前列映射: {current_columns}")
            self.logger.debug(f"列名到索引映射: {column_name_to_index}")
            
            # 重新应用超链接：按列名匹配，而不是位置偏移
            for (orig_row, orig_col), link_info in hyperlinks.items():
                try:
                    # 获取原始列的列名
                    if orig_col not in original_columns:
                        self.logger.warning(f"原始第{orig_col}列没有列名，跳过超链接")
                        continue
                    
                    original_col_name = original_columns[orig_col]
                    
                    # 在当前文件中查找同名列
                    if original_col_name not in column_name_to_index:
                        self.logger.warning(f"在当前文件中找不到列'{original_col_name}'，跳过超链接")
                        continue
                    
                    new_col = column_name_to_index[original_col_name]
                    
                    self.logger.debug(f"超链接重定位: '{original_col_name}' 第{orig_col}列 -> 第{new_col}列")
                    
                    # 应用超链接到新位置
                    if new_col <= ws.max_column:
                        target_cell = ws.cell(row=orig_row, column=new_col)
                        target_cell.hyperlink = link_info['hyperlink']
                        # 保持原有的显示值，不要覆盖
                        if target_cell.value is None:
                            target_cell.value = link_info['value']
                        
                        self.logger.debug(f"超链接已正确应用: 行{orig_row} '{original_col_name}' 第{orig_col}列->第{new_col}列")
                    else:
                        self.logger.warning(f"新列位置 {new_col} 超出范围，跳过超链接")
                        
                except Exception as e:
                    self.logger.warning(f"重新应用超链接失败 (行{orig_row}, 列{orig_col}): {e}")
            
            self.logger.debug("超链接重新应用完成")
            
        except Exception as e:
            self.logger.error(f"处理超链接时发生严重错误: {e}")
            import traceback
            self.logger.error(traceback.format_exc())
    
    def _write_data_to_worksheet(self, ws, result: RankingResult, final_col_map: Dict[str, int], header_row: int):
        """将数据写入工作表"""
        try:
            key_col = ExcelColumns.ORIGINAL_NAME
            
            # 转换数据为字典格式便于查找
            valid_data_dict = result.valid_data.drop_duplicates(subset=[key_col]).set_index(key_col).to_dict('index')
            
            # 写入有效数据
            valid_rows_written = 0
            for row_idx in range(header_row + 1, ws.max_row + 1):
                current_key = ws.cell(row=row_idx, column=final_col_map.get(key_col, 1)).value
                
                if current_key and current_key in valid_data_dict:
                    source_row_data = valid_data_dict[current_key]
                    valid_rows_written += 1
                    
                    # 写入所有列的数据
                    for col_name, col_idx in final_col_map.items():
                        if col_name in source_row_data:
                            value_to_write = source_row_data[col_name]
                            
                            # 处理特殊值
                            if isinstance(value_to_write, str) and value_to_write == "NaN":
                                ws.cell(row=row_idx, column=col_idx).value = "NaN"
                            elif pd.isna(value_to_write):
                                ws.cell(row=row_idx, column=col_idx).value = None
                            else:
                                ws.cell(row=row_idx, column=col_idx).value = value_to_write
            
            # 写入排除的数据（在有效数据下方空两行后）
            if result.excluded_data is not None and not result.excluded_data.empty:
                start_row = header_row + 1 + valid_rows_written + 2
                
                for idx, (_, row_data) in enumerate(result.excluded_data.iterrows()):
                    current_row = start_row + idx
                    
                    for col_name, col_idx in final_col_map.items():
                        if col_name in row_data.index:
                            value_to_write = row_data[col_name]
                            
                            if isinstance(value_to_write, str) and value_to_write == "NaN":
                                ws.cell(row=current_row, column=col_idx).value = "NaN"
                            elif pd.isna(value_to_write):
                                ws.cell(row=current_row, column=col_idx).value = None
                            else:
                                ws.cell(row=current_row, column=col_idx).value = value_to_write
                        else:
                            ws.cell(row=current_row, column=col_idx).value = None
            
            self.logger.info(f"数据写入完成: 有效数据 {valid_rows_written} 行，排除数据 {len(result.excluded_data) if result.excluded_data is not None else 0} 行")
            
        except Exception as e:
            raise ExcelProcessingError(f"写入数据时发生错误: {e}")
    
    def _apply_column_styles(self, ws, final_col_map: Dict[str, int], header_row: int):
        """应用列分组样式"""
        try:
            self.logger.debug("开始应用列分组样式")
            
            if not final_col_map:
                self.logger.warning("列映射为空，跳过样式应用")
                return
            
            # 计算数据范围
            data_start_row = header_row + 1
            data_end_row = ws.max_row
            
            if data_start_row > data_end_row:
                self.logger.warning("数据行范围无效，跳过样式应用")
                return
            
            # 应用每个分组的样式
            for group_name, group_info in self._style_config.items():
                try:
                    columns = group_info["columns"]
                    fill_color = group_info["color"]
                    
                    # 创建填充样式
                    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    
                    # 为该组的每个列应用样式
                    applied_cols = []
                    for col_name in columns:
                        if col_name in final_col_map:
                            col_idx = final_col_map[col_name]
                            
                            # 为该列的所有数据行应用样式
                            for row_idx in range(data_start_row, data_end_row + 1):
                                try:
                                    cell = ws.cell(row=row_idx, column=col_idx)
                                    cell.fill = fill
                                    cell.alignment = self._alignment
                                    cell.border = self._border
                                except Exception as e:
                                    self.logger.warning(f"应用样式到单元格 ({row_idx}, {col_idx}) 失败: {e}")
                            
                            applied_cols.append(col_name)
                    
                    if applied_cols:
                        self.logger.debug(f"已为 '{group_name}' 组的 {len(applied_cols)} 列应用样式")
                        
                except Exception as e:
                    self.logger.warning(f"处理列分组 '{group_name}' 时发生错误: {e}")
            
            self.logger.debug("列分组样式应用完成")
            
        except Exception as e:
            self.logger.warning(f"应用列分组样式时发生错误: {e}")
    
    def _atomic_move_file(self, temp_path: str, final_path: str):
        """原子性地移动文件"""
        try:
            if os.path.exists(final_path):
                backup_path = final_path + '.backup'
                shutil.move(final_path, backup_path)
                try:
                    shutil.move(temp_path, final_path)
                    os.remove(backup_path)
                except Exception as e:
                    # 恢复备份
                    shutil.move(backup_path, final_path)
                    raise
            else:
                shutil.move(temp_path, final_path)
        except Exception as e:
            raise FileOperationError(f"文件移动失败: {e}")
    
    def create_processing_config(self, input_file: str, output_file: str, operation_type: str) -> ProcessingConfig:
        """
        创建处理配置
        
        Args:
            input_file: 输入文件路径
            output_file: 输出文件路径
            operation_type: 操作类型
            
        Returns:
            ProcessingConfig: 处理配置
        """
        return ProcessingConfig(
            input_file=input_file,
            output_file=output_file,
            operation_type=operation_type
        ) 